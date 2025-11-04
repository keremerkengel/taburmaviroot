from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.views.generic import CreateView, UpdateView, DeleteView
from django.utils import timezone
from django.db.models import Count
from django.contrib.auth import logout
from django.shortcuts import redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .forms import RootPlanForm
from .models import RootPlan


def signup_view(request):
    if request.user.is_authenticated:
        return redirect("dashboard")
    if request.method == "POST":
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, "Hesabınız oluşturuldu. Hoş geldiniz!")
            return redirect("dashboard")
        messages.error(request, "Kayıt işlemi başarısız. Lütfen formu kontrol edin.")
    else:
        form = UserCreationForm()
    return render(request, "registration/signup.html", {"form": form})

class OwnerRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        obj = self.get_object()
        return obj.user == self.request.user

class RootPlanCreateView(LoginRequiredMixin, CreateView):
    model = RootPlan
    form_class = RootPlanForm
    template_name = "plans/create_rootplan.html"
    success_url = reverse_lazy("dashboard")

    def form_valid(self, form):
        form.instance.user = self.request.user
        messages.success(self.request, "✅ Ziyaret başarıyla eklendi.")
        return super().form_valid(form)

class RootPlanUpdateView(LoginRequiredMixin, OwnerRequiredMixin, UpdateView):
    model = RootPlan
    form_class = RootPlanForm
    template_name = "plans/edit_rootplan.html"
    success_url = reverse_lazy("dashboard")

    def form_valid(self, form):
        messages.success(self.request, "✏️ Ziyaret başarıyla güncellendi.")
        return super().form_valid(form)

class RootPlanDeleteView(LoginRequiredMixin, OwnerRequiredMixin, DeleteView):
    model = RootPlan
    template_name = "plans/confirm_delete.html"
    success_url = reverse_lazy("dashboard")

    def delete(self, request, *args, **kwargs):
        messages.error(self.request, "❌ Ziyaret kaydı silindi.")
        return super().delete(request, *args, **kwargs)

@login_required
def dashboard(request):
    plans = RootPlan.objects.filter(user=request.user)

    # Zaman bilgisi
    now = timezone.now()
    current_month = now.month
    current_year = now.year

    # Bu ayın ziyaretleri
    monthly_visits = plans.filter(visit_date__year=current_year, visit_date__month=current_month)
    total_visits = monthly_visits.count()

    # Şehir sayısı (farklı iller)
    distinct_cities = monthly_visits.values("city").distinct().count()

    # En son ziyaret
    last_visit = plans.order_by("-visit_date").first()

    context = {
        "plans": plans,
        "now": now,
        "total_visits": total_visits,
        "distinct_cities": distinct_cities,
        "last_visit": last_visit,
    }
    return render(request, "plans/dashboard.html", context)

def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def export_visits_excel(request):
    # Workbook oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Ziyaret Kayıtları"
    
    # Başlık satırı
    headers = ['Tarih', 'İl', 'İlçe', 'Ziyaret Türü', 'Yer Adı', 'Ziyaret Amacı', 'Notlar']
    
    # Başlık stilini ayarla
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    
    # Başlıkları yaz ve stillendir
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Kullanıcının ziyaret kayıtlarını al
    visits = RootPlan.objects.filter(user=request.user).order_by('-visit_date')
    
    # Ziyaret türü çevirileri
    place_type_choices = {
        'klinik': 'Klinik',
        'eczane': 'Eczane', 
        'ciftlik': 'Çiftlik',
        'hastane': 'Hastane',
        'diger': 'Diğer'
    }
    
    # Ziyaret amacı çevirileri  
    visit_purpose_choices = {
        'ilac_tanitim': 'İlaç Tanıtım',
        'tahsil': 'Tahsil',
        'urun_tanitim': 'Ürün Tanıtım',
        'musteri_ziyaret': 'Müşteri Ziyareti',
        'siparis_alma': 'Sipariş Alma',
        'bayi_ziyaret': 'Bayi Ziyareti',
        'pazarlama': 'Pazarlama Faaliyeti',
        'hizmet_sonrasi': 'Hizmet Sonrası Takip',
        'teknik_destek': 'Teknik Destek',
        'fuar_etkinlik': 'Fuar/Etkinlik',
        'egitim_seminer': 'Eğitim/Seminer',
        'yeni_musteri': 'Yeni Müşteri Kazanımı',
        'rekabet_analiz': 'Rekabet Analizi',
        'pazar_arastirma': 'Pazar Araştırması',
        'sikayet_cozum': 'Şikayet Çözümü',
        'kontrol_denetim': 'Kontrol/Denetim',
        'toplanti': 'Toplantı',
        'diger': 'Diğer'
    }
    
    # Verileri yaz
    for row, visit in enumerate(visits, 2):
        # Ziyaret türünü çevir
        place_type_display = place_type_choices.get(visit.place_type, visit.place_type)
        
        # Ziyaret amacını çevir
        visit_purpose_display = visit_purpose_choices.get(visit.visit_purpose, visit.visit_purpose)
        
        row_data = [
            visit.visit_date.strftime('%d.%m.%Y'),
            visit.city,
            visit.district,
            place_type_display,
            visit.place_name,
            visit_purpose_display,
            visit.notes or ''
        ]
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Sütun genişliklerini ayarla
    column_widths = [12, 15, 15, 15, 25, 20, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Response oluştur
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ziyaret_kayitlari.xlsx"'
    
    wb.save(response)
    return response